"""Script de verificación de requisitos del TP5."""
import sys, csv, re
sys.path.insert(0, '.')

from src.validador import validar_fila
from src.filtrador import filtrar, obtener_aps
from src.exportador import exportar_excel
import pandas as pd
import os

SEP = "=" * 60

# ============================================================
# REQ 3: Solo guarda validos; invalidos con motivo
# ============================================================
print(SEP)
print("[REQ 3] Descarte de invalidos con motivo explicito")

filas_prueba = [
    ("ID vacio (invalido)",
     ['', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', 'Wireless-802.11', '2019-02-07', '19:46:08',
      '2019-03-13', '11:27:57', '25', '39517', '505219',
      'DC-9F-DB-12-F3-EA:HCDD', 'DC-BF-E9-1A-B5-D0', 'User-Request'], True),

    ("Tipo_conexion = IP (fila corrupta, invalida)",
     ['603877', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', '192.168.1.1', '2019-02-07', '19:46:08',
      '2019-03-13', '11:27:57', '25', '39517', '505219',
      'DC-9F-DB-12-F3-EA:HCDD', 'DC-BF-E9-1A-B5-D0', 'User-Request'], True),

    ("Fecha en formato DD-MM-YYYY (invalida)",
     ['603877', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', 'Wireless-802.11', '07-02-2019', '19:46:08',
      '2019-03-13', '11:27:57', '25', '39517', '505219',
      'DC-9F-DB-12-F3-EA:HCDD', 'DC-BF-E9-1A-B5-D0', 'User-Request'], True),

    ("Razon vacia + bytes=0 (debe ser VALIDO)",
     ['603877', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', 'Wireless-802.11', '2019-02-07', '19:46:08',
      '2019-03-13', '11:27:57', '25', '0', '0',
      'DC-9F-DB-12-F3-EA:HCDD', 'DC-BF-E9-1A-B5-D0', ''], False),

    ("MAC_AP sin sufijo (invalida)",
     ['603877', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', 'Wireless-802.11', '2019-02-07', '19:46:08',
      '2019-03-13', '11:27:57', '25', '39517', '505219',
      'DC-9F-DB-12-F3-EA', 'DC-BF-E9-1A-B5-D0', 'User-Request'], True),

    ("Solo 10 columnas (invalida)",
     ['603877', '5AA0184E-000001CA', 'd6104707df0cd315', 'invitado-deca',
      '192.168.247.11', 'Wireless-802.11', '2019-02-07', '19:46:08',
      '2019-03-13', '11:27:57'], True),
]

fallos = 0
for desc, fila, debe_invalida in filas_prueba:
    datos, motivo = validar_fila(fila)
    es_invalida = (datos is None)
    ok = (es_invalida == debe_invalida)
    if not ok:
        fallos += 1
    icono = "OK  " if ok else "FAIL"
    detalle = ("invalido: " + motivo[:60]) if es_invalida else "valido"
    print(f"  {icono}  {desc}")
    print(f"         -> {detalle}")

print(f"  STATUS: {'OK' if fallos == 0 else 'FALLO en ' + str(fallos) + ' casos'}")

# ============================================================
# REQ 4: Filtrado por AP + fechas
# ============================================================
print()
print(SEP)
print("[REQ 4] Filtrado por AP y rango de fechas")

validos = []
with open('export-2019-to-now-v4.csv', encoding='utf-8', errors='replace') as f:
    reader = csv.reader(f)
    next(reader)
    for i, fila in enumerate(reader):
        if i >= 5000:
            break
        datos, _ = validar_fila(fila)
        if datos:
            validos.append(datos)

df = pd.DataFrame(validos)
aps = obtener_aps(df)
print(f"  -> APs disponibles: {len(aps)}")

ap_test = aps[0]
df_res = filtrar(df, ap_test, '2019-01-01', '2019-06-30')
print(f"  -> AP: {ap_test}")
print(f"  -> Rango: 2019-01-01 a 2019-06-30")
print(f"  -> Registros encontrados: {len(df_res)}")

fuera = df_res[
    (df_res['Inicio_Dia'] < '2019-01-01') | (df_res['Inicio_Dia'] > '2019-06-30')
]
print(f"  -> Registros fuera de rango (deben ser 0): {len(fuera)}")
ap_incorrecto = df_res[df_res['MAC_AP'] != ap_test]
print(f"  -> Registros con AP incorrecto (deben ser 0): {len(ap_incorrecto)}")
ok4 = (len(fuera) == 0 and len(ap_incorrecto) == 0)
print(f"  STATUS: {'OK' if ok4 else 'FALLO'}")

# ============================================================
# REQ 5: Exportacion a Excel
# ============================================================
print()
print(SEP)
print("[REQ 5] Exportacion a Excel con openpyxl")

ruta_test = "_verificacion_test.xlsx"
ruta = exportar_excel(df_res, ap_test, '2019-01-01', '2019-06-30', ruta_test)
existe = os.path.isfile(ruta)
tamano = os.path.getsize(ruta) if existe else 0

import openpyxl
wb = openpyxl.load_workbook(ruta)
hojas = wb.sheetnames
print(f"  -> Archivo generado: {existe}")
print(f"  -> Tamanio: {tamano:,} bytes")
print(f"  -> Hojas: {hojas}")
ws = wb["Conexiones Filtradas"]
filas_excel = ws.max_row
print(f"  -> Filas en hoja principal: {filas_excel} (datos + 4 cabecera)")
print(f"  STATUS: {'OK' if existe and tamano > 0 else 'FALLO'}")
os.remove(ruta)

# ============================================================
# REQ 6: re es central (no decorativo)
# ============================================================
print()
print(SEP)
print("[REQ 6] Verificacion de uso de re en el codigo fuente")

for modulo in ["src/patrones.py", "src/validador.py", "main.py"]:
    with open(modulo, encoding="utf-8") as f:
        contenido = f.read()
    count_re_compile  = contenido.count("re.compile")
    count_fullmatch   = contenido.count("fullmatch")
    count_import_re   = contenido.count("import re")
    print(f"  {modulo}:")
    print(f"    import re: {count_import_re}, re.compile: {count_re_compile}, fullmatch: {count_fullmatch}")

print()
print(SEP)
print("VERIFICACION COMPLETA")
print(SEP)
