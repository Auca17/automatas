# Exportación de resultados a Excel sin formato ni colores.

import openpyxl
from datetime import datetime
import pandas as pd
import os

# Definición de columnas a exportar (campo_interno → título en Excel)
COLUMNAS_EXPORT: list[tuple[str, str]] = [
    ("Usuario",       "Usuario"),
    ("MAC_Cliente",   "MAC Cliente"),
    ("IP_NAS_AP",     "IP del AP"),
    ("Inicio_Dia",    "Inicio Día"),
    ("Inicio_Hora",   "Inicio Hora"),
    ("Fin_Dia",       "Fin Día"),
    ("Fin_Hora",      "Fin Hora"),
    ("Session_Time",  "Duración (seg)"),
    ("Input_Octects", "Entrada (bytes)"),
    ("Output_Octects","Salida (bytes)"),
    ("Razon",         "Razón Terminación"),
    ("ID",            "ID Registro"),
    ("ID_Sesion",     "ID Sesión"),
    ("MAC_AP",        "MAC AP"),
]

def exportar_excel(
    df_resultado: pd.DataFrame,
    mac_ap: str,
    fecha_desde: str,
    fecha_hasta: str,
    ruta: str | None = None,
) -> str:
    """
    Exporta el DataFrame filtrado a un archivo .xlsx sin formato.
    """
    if ruta is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"conexiones_{ts}.xlsx"
        ruta = os.path.join(os.getcwd(), nombre)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Conexiones Filtradas")

    # FILA 1 — Información del filtro (texto plano)
    ws.append([
        f"AP: {mac_ap}",
        f"Período: {fecha_desde} → {fecha_hasta}",
        f"Registros: {len(df_resultado):,}",
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ])

    # FILA 2 — Encabezados de columnas
    encabezados = [col[1] for col in COLUMNAS_EXPORT]
    ws.append(encabezados)

    # FILAS DE DATOS
    campos = [col[0] for col in COLUMNAS_EXPORT]
    datos = df_resultado[campos].astype(str).values.tolist()
    for fila in datos:
        ws.append(fila)

    wb.save(ruta)
    return ruta

def exportar_invalidos_excel(
    invalidos: list[dict],
    ruta: str | None = None,
) -> str:
    
    if ruta is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta = os.path.join(os.getcwd(), f"invalidos_{ts}.xlsx")

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Registros Inválidos")

    # Encabezados
    ws.append(["N° Fila CSV", "ID del Registro", "Motivo del Descarte"])

    # Datos
    for inv in invalidos:
        ws.append([
            str(inv.get("num_fila", "")),
            str(inv.get("ID", "")),
            str(inv.get("motivo", ""))
        ])

    wb.save(ruta)
    return ruta
